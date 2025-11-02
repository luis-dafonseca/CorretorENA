#------------------------------------------------------------------------------#
'''Wrapper to OpenPyXL

def read_names_from_spreadsheet(fname: str, first_cell: str) -> list[str]
    Open spreadsheet, read names and close it

ResultsSheet: Class to wrap the spreadsheet to store and save the results

    def __init__(self) -> None
        Initialize class instance

    def write_names(self, names: list[str]) -> None
        Write names to spreadsheet

    def get_name(self, ii: int) -> str
        Return a candidate name from spreadsheet

    def add_grade(self, ii: int, answers: Answers) -> None
        Write grade and status of candidate to spreadsheet

    def save(self, fname: str) -> None
        Save spreadsheet to file
'''

from openpyxl            import Workbook, load_workbook
from openpyxl.styles     import Alignment, Font
from openpyxl.utils.cell import (column_index_from_string,
                                 coordinate_from_string)

from grading.answers import Answers

#------------------------------------------------------------------------------#
def read_names_from_spreadsheet(fname: str, first_cell: str) -> list[str]:
    '''Read names from spreadsheet'''

    xy = coordinate_from_string(first_cell)
    cc = column_index_from_string(xy[0]) - 1
    ll = xy[1]

    input = load_workbook(fname)
    sheet = input.active

    names = [row[cc]
             for row in sheet.iter_rows(min_row=ll, values_only=True)]

    input.close()

    return names

#------------------------------------------------------------------------------#
class ResultsSheet:
    '''Class to wrap the spreadsheet to store and save the results'''

    #--------------------------------------------------------------------------#
    def __init__(self) -> None:
        '''Initialize class instance'''

        self.has_names = False

        self.wb    = Workbook()
        self.sheet = self.wb.active

        self.sheet['A1'] = 'Nome'
        self.sheet['B1'] = 'Nota'
        self.sheet['C1'] = 'Resultado'

        for nn in range(1, 31):
            cell = self.sheet.cell(row=1, column=nn+3)
            cell.value = nn

        for nn in range(1, 34):
            cell = self.sheet.cell(row=1, column=nn)
            cell.alignment = Alignment(horizontal='center')
            cell.font      = Font(bold=True)

        self.sheet.column_dimensions['A'].width = 16
        self.sheet.column_dimensions['C'].width = 12

        # Summary
        self.total      = 0
        self.eliminated = 0
        self.absent     = 0
        self.approved   = 0
        self.reproved   = 0

    #--------------------------------------------------------------------------#
    def write_names(self, names: list[str]) -> None:
        '''Write names to spreadsheet'''

        max_chars = 10

        for rr, name in enumerate(names):

            name = name.strip()

            max_chars = max(max_chars, len(name))

            self.sheet.cell(row=rr+2, column=1).value = name

        self.sheet.column_dimensions['A'].width = max_chars * 1.6

        self.has_names = True

    #--------------------------------------------------------------------------#
    def get_name(self, ii: int) -> str:
        '''Return a candidate name from spreadsheet'''

        if self.has_names:
            return str(self.sheet.cell(row=ii+2, column=1).value)
        else:
            return str(f'Candidato {ii+1}')

    #--------------------------------------------------------------------------#
    def add_grade(self, ii: int, answers: Answers) -> None:
        '''Write grade and status of candidate to spreadsheet'''

        self.total += 1

        aa = self.sheet.cell(ii+2, 1)
        bb = self.sheet.cell(ii+2, 2)
        cc = self.sheet.cell(ii+2, 3)

        if not self.has_names:
            aa.value = str(f'Candidato {ii+1}')

        bb.value     = answers.get_score()
        bb.alignment = Alignment(horizontal='center')

        if answers.is_eliminated():
            cc.value = 'Eliminado'
            self.eliminated += 1

        elif answers.is_absent():
            cc.value = 'Ausente'
            self.absent += 1

        elif answers.is_approved():
            cc.value = 'Aprovado'
            self.approved += 1

        else:
            cc.value = 'Reprovado'
            self.reproved += 1

        for nn in range(30):
            cell = self.sheet.cell(row=ii+2, column=nn+4)
            cell.value = answers.correct[nn]
            cell.alignment = Alignment(horizontal='center')

    #--------------------------------------------------------------------------#
    def summary(self) -> dict[str, int]:
        '''Return a results summary'''

        return {
            'total':      self.total,
            'eliminated': self.eliminated,
            'absent':     self.absent,
            'approved':   self.approved,
            'reproved':   self.reproved
        }

    #--------------------------------------------------------------------------#
    def save(self, fname: str) -> None:
        '''Save spreadsheet to file'''

        self.wb.save(fname)

#------------------------------------------------------------------------------#

