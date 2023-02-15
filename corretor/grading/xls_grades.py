#------------------------------------------------------------------------------#
'''Wraper to OpenPyXL'''

from openpyxl            import Workbook, load_workbook
from openpyxl.styles     import Alignment, Font
from openpyxl.utils.cell import (column_index_from_string,
                                 coordinate_from_string)

from grading.answers import Answers

#------------------------------------------------------------------------------#
class XLSGrades:
    '''Class to wrap the spreadsheet to store and save the results'''

    #--------------------------------------------------------------------------#
    def __init__(self, fname: str) -> None:
        '''Initialize class instance'''

        self.fname = fname

        self.wb    = Workbook()
        self.sheet = self.wb.active

        self.sheet['A1'] = 'Nome'
        self.sheet['B1'] = 'Nota'
        self.sheet['C1'] = 'Status'

        self.sheet['A1'].alignment = Alignment(horizontal='center')
        self.sheet['B1'].alignment = Alignment(horizontal='center')
        self.sheet['C1'].alignment = Alignment(horizontal='center')

        self.sheet['A1'].font = Font(bold=True)
        self.sheet['B1'].font = Font(bold=True)
        self.sheet['C1'].font = Font(bold=True)

        self.sheet.column_dimensions['A'].width = 16
        self.sheet.column_dimensions['C'].width = 12

        self.has_names = False

    #--------------------------------------------------------------------------#
    def write_names(self, names: list[str]) -> None:
        '''Write names to spreadsheet'''

        ww = 4

        for rr, name in enumerate(names):
            name = name.strip()
            ww = max(ww, len(name))
            self.sheet.cell(row=rr+2, column=1).value = name

        self.sheet.column_dimensions['A'].width = ww * 1.6

        self.has_names = True

    #--------------------------------------------------------------------------#
    def read_names(self, fname: str, first_name: str) -> None:
        '''Copy names to spreadsheet'''

        xls   = load_workbook(fname)
        sheet = xls.active

        xy = coordinate_from_string(first_name)
        cc = column_index_from_string(xy[0]) - 1
        ll = xy[1]

        for rr, row in enumerate(sheet.iter_rows(min_row=ll, values_only=True)):
            self.sheet.cell(row=rr+2,column=1).value = row[cc]

        xls.close()

        self.has_names = True

    #--------------------------------------------------------------------------#
    def get_name(self, ii: int) -> str:
        '''Return a candidate name from spreadsheet'''

        if self.has_names:
            return str(self.sheet.cell(row=ii+2, column=1).value)
        else:
            return str(f'Candidato {ii+1}')

    #--------------------------------------------------------------------------#
    def add_grade(self, ii: int, answers: Answers) -> None:
        '''Write grade and status of candidate to spreadsheet'''

        aa = self.sheet.cell(ii+2, 1)
        bb = self.sheet.cell(ii+2, 2)
        cc = self.sheet.cell(ii+2, 3)

        if not self.has_names:
            aa.value = str(f'Candidato {ii+1}')

        bb.value     = answers.get_score()
        bb.alignment = Alignment(horizontal='center')

        if   answers.is_eliminated(): cc.value = 'Eliminado'
        elif answers.is_absent    (): cc.value = 'Ausente'
        elif answers.is_approved  (): cc.value = 'Aprovado'
        else:                         cc.value = 'Reprovado'

    #--------------------------------------------------------------------------#
    def reset(self) -> None:
        '''Remove all names, grades and status'''

        for row in self.sheet.iter_rows(min_row=1):
            for cell in row:
                cell.value = None

    #--------------------------------------------------------------------------#
    def save(self) -> None:
        '''Save spreadsheet'''

        self.wb.save(self.fname)

#------------------------------------------------------------------------------#

