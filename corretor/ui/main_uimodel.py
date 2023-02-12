#------------------------------------------------------------------------------#

import fitz
import tempfile
from pathlib import Path

from openpyxl            import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from grading.grade_exam import grade_exam
from grading.xls_grades import XLSGrades
from grading.ena_form   import ENAForm

import grading.rectangles as rects

from ui.keys_model import KeysModel

#------------------------------------------------------------------------------#
class MainUIModel:

    #--------------------------------------------------------------------------#
    def __init__(self) -> None:

        self.has_model       = False
        self.has_answers     = False
        self.has_annotations = False
        self.has_grades      = False
        self.has_names       = False

        self.names = []

        self.model       = None
        self.answers     = None
        self.annotations = None
        self.xls_grades  = None

        self.num_answers = 0
        self.num_names   = 0

        self.answers_fname = ''

        self.keys_model = KeysModel()

    #--------------------------------------------------------------------------#
    def __del__(self) -> None:

        if self.has_model:
            self.model.close()

        if self.has_answers:
            self.answers.close()

    #--------------------------------------------------------------------------#
    def run(self, progress) -> bool:

        if self.has_names:
            self.xls_grades.write_names(self.names)

        self.annotations = fitz.open()

        finished = grade_exam(
            self.model,
            self.keys_model.keys,
            self.answers,
            self.annotations,
            self.xls_grades,
            progress
        )

        if finished:
            self.annotations.save(self.fname_annotations)
            self.xls_grades .save()
        else:
            self.xls_grades.reset()

        return finished

    #--------------------------------------------------------------------------#
    def ready_to_run(self) -> bool:
        return (
            self.has_model       and
            self.has_answers     and
            self.has_annotations and
            self.has_grades
        )

    #--------------------------------------------------------------------------#
    def set_model(self, fname) -> None:

        new_model = fitz.open(fname)

        nn   = new_model.page_count
        name = Path(fname).name

        if nn == 0:
            raise ValueError((
                f'O arquivo {name} não é um modelo!\n\n'
                'Ele não contém nenhuma página.\n'
            ))
        elif nn > 1:
            raise ValueError((
                f'O arquivo {name} não é um modelo!\n\n'
                'Ele contém mais do que uma página.\n'
            ))

        if self.has_model:
            self.model.close()

        self.model     = new_model
        self.has_model = True

    #--------------------------------------------------------------------------#
    def set_answers(self, fname: str) -> None:

        fpath = Path(fname).resolve()

        new_answers = fitz.open(fname)

        nn = new_answers.page_count

        if nn == 0:
            raise ValueError( f'O arquivo {fpath.name} não contém nenhuma página.\n' )

        if self.has_answers:
            self.answers.close()

        self.num_answers   = nn
        self.answers       = new_answers
        self.answers_fname = str(fname)
        self.has_answers   = True

        if self.has_names and (self.num_answers != self.num_names):
            raise IndexError((
                f'Cuidado: A quantidade de respostas ({self.num_answers})'
                f'não coincide com a quantidade de nomes ({self.num_names}).\n'
            ))

    #--------------------------------------------------------------------------#
    def set_names(self, fname:str, first_name:str) -> None:

        xls = load_workbook(fname)

        fpath = Path(fname).resolve()

        try:
            sheet = xls.active

            xy = coordinate_from_string(first_name)
            cc = column_index_from_string(xy[0]) - 1
            ll = xy[1]

            self.names = []

            for row in sheet.iter_rows(min_row=ll, values_only=True):
                self.names.append(str(row[cc]).strip().upper())

            xls.close()

        except IOError as er:
            self.names = []
            raise IOError(
                f'Não foi possível ler os nomes do arquivo {fpath.name}!\n\n{str(er)}'
            )

        self.num_names = len(self.names)
        self.has_names = True

        if self.has_answers and (self.num_answers != self.num_names):
            raise IndexError((
                f'Cuidado: A quantidade de respostas ({self.num_answers})'
                f' não coincide com a quantidade de nomes ({self.num_names}).\n'
            ))

    #--------------------------------------------------------------------------#
    def remove_names(self) -> None:
        self.names     = []
        self.num_names = 0
        self.has_names = False

    #--------------------------------------------------------------------------#
    def set_annotations(self, fname: str) -> None:

        self.fname_annotations = fname
        self.has_annotations   = True

    #--------------------------------------------------------------------------#
    def set_grades(self, fname: str) -> None:

        self.xls_grades = XLSGrades(fname)
        self.has_grades = True

    #--------------------------------------------------------------------------#
    def get_model_pdf(self) -> str:

        pixmap = self.model[0].get_pixmap(dpi=rects.DPI, colorspace='GRAY')

        new_pdf = fitz.open()

        page = new_pdf.new_page(
            width  = rects.PAGE.width,
            height = rects.PAGE.height
        )

        form = ENAForm(page)
        form.insert_pixmap(pixmap)
        form.insert_rects()
        form.commit()

        self.temp_file = tempfile.NamedTemporaryFile(
            prefix='modelo_',
            suffix='.pdf'
        )
        new_pdf.save(self.temp_file.name)

        return self.temp_file.name

    #--------------------------------------------------------------------------#
    def get_keys_pdf(self) -> str:

        pixmap = self.model[0].get_pixmap(dpi=rects.DPI, colorspace='GRAY')

        new_pdf = fitz.open()

        page = new_pdf.new_page(
            width  = rects.PAGE.width,
            height = rects.PAGE.height
        )

        form = ENAForm(page)
        form.insert_pixmap(pixmap)
        form.insert_keys(self.keys_model.keys)
        form.commit()

        self.temp_file = tempfile.NamedTemporaryFile(
            prefix = 'gabarito_',
            suffix = '.pdf'
        )
        new_pdf.save( self.temp_file.name )

        return self.temp_file.name

    #--------------------------------------------------------------------------#
    def get_answers_pdf(self) -> str:
        return self.answers_fname

    #--------------------------------------------------------------------------#
    def has_conflict(self) -> bool:

        return (
            self.has_answers and
            self.has_names   and
            self.num_names != self.num_answers
        )

#------------------------------------------------------------------------------#
